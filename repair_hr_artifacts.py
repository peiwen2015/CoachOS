#!/usr/bin/env python3
from __future__ import annotations

import argparse
import re
import sqlite3
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.cell_range import CellRange
from garmin_fit_sdk import Decoder, Stream


ROOT = Path(__file__).resolve().parent
DEFAULT_EXCEL_DIR = ROOT / "EXCEL"
DEFAULT_FIT_DIR = ROOT / "FIT"
DEFAULT_DB_PATH = ROOT / "analysis_platform" / "running_analytics.sqlite"

ACTIVITY_ID_PATTERN = re.compile(r"_(\d+)\.xlsx$")
FIT_ACTIVITY_ID_PATTERN = re.compile(r"(\d{10,})")


def decode_fit(path: Path):
    stream = Stream.from_file(str(path))
    messages, errors = Decoder(stream).read()
    if errors:
        raise RuntimeError(f"FIT decode errors: {errors}")
    return messages


def backfill_sqlite_activity_max_hr(db_path: Path, month: str | None = None) -> int:
    params: tuple[str, ...] = ()
    sql = f"""
    WITH split_max AS (
      SELECT activity_id, MAX(max_hr) AS split_max_hr
      FROM kilometer_split
      GROUP BY activity_id
    )
    UPDATE activity
    SET max_hr = (
      SELECT split_max_hr
      FROM split_max
      WHERE split_max.activity_id = activity.id
    )
    WHERE id IN (
      SELECT activity_id
      FROM split_max
      WHERE split_max_hr IS NOT NULL
    );
    """
    if month:
        params = (f"{month}-%",)
        sql = f"""
        WITH split_max AS (
          SELECT ks.activity_id, MAX(ks.max_hr) AS split_max_hr
          FROM kilometer_split ks
          JOIN activity a ON a.id = ks.activity_id
          WHERE a.activity_start_time LIKE ?
          GROUP BY ks.activity_id
        )
        UPDATE activity
        SET max_hr = (
          SELECT split_max_hr
          FROM split_max
          WHERE split_max.activity_id = activity.id
        )
        WHERE id IN (
          SELECT activity_id
          FROM split_max
          WHERE split_max_hr IS NOT NULL
        );
        """
    with sqlite3.connect(db_path) as connection:
        cursor = connection.cursor()
        cursor.execute(sql, params)
        connection.commit()
        return connection.total_changes


def load_fit_activity_avg_power_by_garmin_id(fit_dir: Path, month: str | None = None) -> dict[str, int]:
    mapping: dict[str, int] = {}
    if not fit_dir.exists():
        return mapping
    month_dir = fit_dir / month if month else fit_dir
    if not month_dir.exists():
        return mapping
    for fit_path in sorted(month_dir.rglob("*.fit")):
        match = FIT_ACTIVITY_ID_PATTERN.search(fit_path.stem)
        if not match:
            continue
        try:
            messages = decode_fit(fit_path)
        except Exception:
            continue
        sessions = messages.get("session_mesgs", []) or []
        session = sessions[0] if sessions else {}
        avg_power = session.get("avg_power")
        if isinstance(avg_power, (int, float)):
            mapping[match.group(1)] = int(round(float(avg_power)))
    return mapping


def backfill_sqlite_activity_avg_power(db_path: Path, fit_dir: Path, month: str | None = None) -> int:
    avg_power_by_garmin_id = load_fit_activity_avg_power_by_garmin_id(fit_dir, month)
    if not avg_power_by_garmin_id:
        return 0

    params: tuple[str, ...] = ()
    query = """
        SELECT id, garmin_activity_id
        FROM activity
        WHERE garmin_activity_id IS NOT NULL
          AND avg_power_w IS NULL
    """
    if month:
        query += " AND activity_start_time LIKE ?"
        params = (f"{month}-%",)

    updates = 0
    with sqlite3.connect(db_path) as connection:
        rows = connection.execute(query, params).fetchall()
        for row in rows:
            garmin_activity_id = str(row[1])
            avg_power = avg_power_by_garmin_id.get(garmin_activity_id)
            if avg_power is None:
                continue
            cursor = connection.execute(
                """
                UPDATE activity
                SET avg_power_w = ?, updated_at = CURRENT_TIMESTAMP
                WHERE id = ?
                """,
                (avg_power, row[0]),
            )
            updates += cursor.rowcount or 0
        connection.commit()
    return updates


def load_fit_activity_name_by_garmin_id(fit_dir: Path, month: str | None = None) -> dict[str, str]:
    mapping: dict[str, str] = {}
    if not fit_dir.exists():
        return mapping
    month_dir = fit_dir / month if month else fit_dir
    if not month_dir.exists():
        return mapping
    for fit_path in sorted(month_dir.rglob("*.fit")):
        match = FIT_ACTIVITY_ID_PATTERN.search(fit_path.stem)
        if not match:
            continue
        try:
            messages = decode_fit(fit_path)
        except Exception:
            continue
        sessions = messages.get("session_mesgs", []) or []
        session = sessions[0] if sessions else {}
        workout_messages = messages.get("workout_mesgs", []) or []
        workout = workout_messages[0] if workout_messages else {}
        activity_name = session.get("name") or session.get("activity_name") or workout.get("wkt_name")
        if isinstance(activity_name, (list, tuple)):
            for item in activity_name:
                text = str(item or "").strip()
                if text:
                    activity_name = text
                    break
            else:
                activity_name = None
        if activity_name not in ("", None):
            mapping[match.group(1)] = str(activity_name)
    return mapping


def backfill_sqlite_activity_name(db_path: Path, fit_dir: Path, month: str | None = None) -> int:
    activity_name_by_garmin_id = load_fit_activity_name_by_garmin_id(fit_dir, month)
    if not activity_name_by_garmin_id:
        return 0

    params: tuple[str, ...] = ()
    query = """
        SELECT id, garmin_activity_id
        FROM activity
        WHERE garmin_activity_id IS NOT NULL
          AND (activity_name IS NULL OR TRIM(activity_name) = '')
    """
    if month:
        query += " AND activity_start_time LIKE ?"
        params = (f"{month}-%",)

    updates = 0
    with sqlite3.connect(db_path) as connection:
        rows = connection.execute(query, params).fetchall()
        for row in rows:
            garmin_activity_id = str(row[1])
            activity_name = activity_name_by_garmin_id.get(garmin_activity_id)
            if activity_name is None:
                continue
            cursor = connection.execute(
                """
                UPDATE activity
                SET activity_name = ?, updated_at = CURRENT_TIMESTAMP
                WHERE id = ?
                """,
                (activity_name, row[0]),
            )
            updates += cursor.rowcount or 0
        connection.commit()
    return updates


def load_activity_max_hr_by_garmin_id(db_path: Path, month: str | None = None) -> dict[str, int]:
    query = """
        SELECT garmin_activity_id, max_hr
        FROM activity
        WHERE garmin_activity_id IS NOT NULL
          AND max_hr IS NOT NULL
    """
    params: tuple[str, ...] = ()
    if month:
        query += " AND activity_start_time LIKE ?"
        params = (f"{month}-%",)
    with sqlite3.connect(db_path) as connection:
        rows = connection.execute(query, params).fetchall()
    return {str(garmin_activity_id): int(max_hr) for garmin_activity_id, max_hr in rows}


def load_activity_avg_power_by_garmin_id(db_path: Path, month: str | None = None) -> dict[str, int]:
    query = """
        SELECT garmin_activity_id, avg_power_w
        FROM activity
        WHERE garmin_activity_id IS NOT NULL
          AND avg_power_w IS NOT NULL
    """
    params: tuple[str, ...] = ()
    if month:
        query += " AND activity_start_time LIKE ?"
        params = (f"{month}-%",)
    with sqlite3.connect(db_path) as connection:
        rows = connection.execute(query, params).fetchall()
    return {str(garmin_activity_id): int(avg_power_w) for garmin_activity_id, avg_power_w in rows}


def parse_garmin_activity_id(path: Path) -> str | None:
    match = ACTIVITY_ID_PATTERN.search(path.name)
    return match.group(1) if match else None


def clone_row_style(ws, source_row: int, target_row: int) -> None:
    for col in range(1, 3):
        source = ws.cell(source_row, col)
        target = ws.cell(target_row, col)
        if source.has_style:
            target._style = copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
        if source.alignment:
            target.alignment = copy(source.alignment)
        if source.font:
            target.font = copy(source.font)
        if source.fill:
            target.fill = copy(source.fill)
        if source.border:
            target.border = copy(source.border)


def rebuild_activity_info_sheet_with_inserted_row(workbook, insert_at: int, label: str, value: str, style_source_row: int) -> None:
    old = workbook["活動資訊"]
    new = workbook.create_sheet("活動資訊__tmp", workbook.sheetnames.index("活動資訊"))

    for column_letter, dimension in old.column_dimensions.items():
        new.column_dimensions[column_letter].width = dimension.width
    for row_index, dimension in old.row_dimensions.items():
        target_row = row_index + 1 if row_index >= insert_at else row_index
        new.row_dimensions[target_row].height = dimension.height

    for row in old.iter_rows():
        for cell in row:
            target_row = cell.row + 1 if cell.row >= insert_at else cell.row
            target = new.cell(target_row, cell.column, cell.value)
            if cell.has_style:
                target._style = copy(cell._style)
            if cell.number_format:
                target.number_format = cell.number_format
            if cell.alignment:
                target.alignment = copy(cell.alignment)
            if cell.font:
                target.font = copy(cell.font)
            if cell.fill:
                target.fill = copy(cell.fill)
            if cell.border:
                target.border = copy(cell.border)
            if cell.protection:
                target.protection = copy(cell.protection)

    for merged in list(old.merged_cells.ranges):
        merged_range = CellRange(str(merged))
        if merged_range.min_row >= insert_at:
            merged_range.shift(row_shift=1, col_shift=0)
        elif merged_range.max_row >= insert_at:
            merged_range.shift(row_shift=1, col_shift=0)
        new.merge_cells(str(merged_range))

    for col in range(1, 3):
        source = old.cell(style_source_row, col)
        target = new.cell(insert_at, col)
        if source.has_style:
            target._style = copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
        if source.alignment:
            target.alignment = copy(source.alignment)
        if source.font:
            target.font = copy(source.font)
        if source.fill:
            target.fill = copy(source.fill)
        if source.border:
            target.border = copy(source.border)
        if source.protection:
            target.protection = copy(source.protection)

    new.cell(insert_at, 1, label)
    new.cell(insert_at, 2, value)

    workbook.remove(old)
    new.title = "活動資訊"


def ensure_activity_hr_row(workbook_path: Path, activity_max_hr: int) -> bool:
    wb = load_workbook(workbook_path)
    if "活動資訊" not in wb.sheetnames:
        return False

    ws = wb["活動資訊"]
    personal_row = None
    activity_row = None
    changed = False

    for row in range(1, ws.max_row + 1):
        label = ws.cell(row, 1).value
        if label == "最大心率":
            ws.cell(row, 1, "個人最大心率")
            changed = True
            personal_row = row
        elif label == "個人最大心率":
            personal_row = row
        elif label == "活動最高心率":
            activity_row = row
        elif label == "Critical Power (W)":
            ws.cell(row, 1, "個人臨界功率")
            changed = True
        elif label == "Critical Power(W)":
            ws.cell(row, 1, "個人臨界功率")
            changed = True
        elif label == "Training Effect (Aerobic)":
            ws.cell(row, 1, "有氧訓練效果")
            changed = True
        elif label == "Training Effect (Anaerobic)":
            ws.cell(row, 1, "無氧訓練效果")
            changed = True
        elif label == "Training Load":
            ws.cell(row, 1, "訓練負荷")
            changed = True
        elif label == "Recovery Time (hr)":
            ws.cell(row, 1, "恢復時間（小時）")
            changed = True
        elif label == "Stamina 起始 (%)":
            ws.cell(row, 1, "體力起始 (%)")
            changed = True
        elif label == "Stamina 結束 (%)":
            ws.cell(row, 1, "體力結束 (%)")
            changed = True

    if personal_row is None:
        return False

    if activity_row is None:
        insert_at = personal_row + 1
        ws.insert_rows(insert_at)
        clone_row_style(ws, personal_row, insert_at)
        ws.cell(insert_at, 1, "活動最高心率")
        ws.cell(insert_at, 2, activity_max_hr)
        changed = True
    else:
        if ws.cell(activity_row, 2).value != activity_max_hr:
            ws.cell(activity_row, 2, activity_max_hr)
            changed = True

    if changed or ws.cell(personal_row, 1).value != "個人最大心率":
        wb.save(workbook_path)
        return True
    return False


def ensure_activity_avg_power_row(workbook_path: Path, activity_avg_power: int) -> bool:
    wb = load_workbook(workbook_path)
    if "活動資訊" not in wb.sheetnames:
        return False

    ws = wb["活動資訊"]
    pace_row = None
    power_row = None
    changed = False

    for row in range(1, ws.max_row + 1):
        label = ws.cell(row, 1).value
        if label == "平均配速":
            pace_row = row
        elif label in ("平均功率", "平均功率 (W)"):
            power_row = row

    if pace_row is None:
        return False

    if power_row is None:
        insert_at = pace_row + 1
        rebuild_activity_info_sheet_with_inserted_row(wb, insert_at, "平均功率", f"{activity_avg_power} W", pace_row)
        changed = True
    else:
        if ws.cell(power_row, 1).value != "平均功率":
            ws.cell(power_row, 1, "平均功率")
            changed = True
        if ws.cell(power_row, 2).value != f"{activity_avg_power} W":
            ws.cell(power_row, 2, f"{activity_avg_power} W")
            changed = True

    if changed:
        wb.save(workbook_path)
        return True
    return False


def ensure_activity_name_row(workbook_path: Path, activity_name: str) -> bool:
    wb = load_workbook(workbook_path)
    if "活動資訊" not in wb.sheetnames:
        return False

    ws = wb["活動資訊"]
    changed = False

    for row in range(1, ws.max_row + 1):
        if ws.cell(row, 1).value != "活動名稱":
            continue
        current_value = ws.cell(row, 2).value
        if current_value not in ("", None):
            return False
        ws.cell(row, 2, activity_name)
        changed = True
        break

    if changed:
        wb.save(workbook_path)
        return True
    return False


def repair_excel_workbooks(excel_dir: Path, db_path: Path, fit_dir: Path, month: str | None = None) -> tuple[int, int]:
    hr_by_garmin_id = load_activity_max_hr_by_garmin_id(db_path, month)
    avg_power_by_garmin_id = load_activity_avg_power_by_garmin_id(db_path, month)
    activity_name_by_garmin_id = load_fit_activity_name_by_garmin_id(fit_dir, month)
    target_dir = excel_dir / month if month else excel_dir
    scanned = 0
    updated = 0
    for workbook_path in target_dir.rglob("*.xlsx"):
        garmin_activity_id = parse_garmin_activity_id(workbook_path)
        if not garmin_activity_id:
            continue
        activity_max_hr = hr_by_garmin_id.get(garmin_activity_id)
        activity_avg_power = avg_power_by_garmin_id.get(garmin_activity_id)
        activity_name = activity_name_by_garmin_id.get(garmin_activity_id)
        if activity_max_hr is None and activity_avg_power is None and activity_name is None:
            continue
        scanned += 1
        workbook_updated = False
        if activity_max_hr is not None and ensure_activity_hr_row(workbook_path, activity_max_hr):
            workbook_updated = True
        if activity_name is not None and ensure_activity_name_row(workbook_path, activity_name):
            workbook_updated = True
        if activity_avg_power is not None and ensure_activity_avg_power_row(workbook_path, activity_avg_power):
            workbook_updated = True
        if workbook_updated:
            updated += 1
    return scanned, updated


def main() -> None:
    parser = argparse.ArgumentParser(description="Repair activity name, max heart rate, and average power artifacts in SQLite and Excel workbooks.")
    parser.add_argument("--sqlite-db", type=Path, default=DEFAULT_DB_PATH)
    parser.add_argument("--excel-dir", type=Path, default=DEFAULT_EXCEL_DIR)
    parser.add_argument("--fit-dir", type=Path, default=DEFAULT_FIT_DIR)
    parser.add_argument("--month", help="Repair only one month, e.g. 2026-07")
    parser.add_argument("--skip-sqlite", action="store_true")
    parser.add_argument("--skip-excel", action="store_true")
    args = parser.parse_args()

    if not args.skip_sqlite:
        sqlite_changes = backfill_sqlite_activity_max_hr(args.sqlite_db, args.month)
        sqlite_name_changes = backfill_sqlite_activity_name(args.sqlite_db, args.fit_dir, args.month)
        sqlite_power_changes = backfill_sqlite_activity_avg_power(args.sqlite_db, args.fit_dir, args.month)
        print(
            "SQLite repaired: "
            f"activity_name {sqlite_name_changes} row updates, "
            f"max_hr {sqlite_changes} row updates, "
            f"avg_power {sqlite_power_changes} row updates"
        )

    if not args.skip_excel:
        scanned, updated = repair_excel_workbooks(args.excel_dir, args.sqlite_db, args.fit_dir, args.month)
        print(f"Excel repaired: {updated}/{scanned} workbooks updated")


if __name__ == "__main__":
    main()
